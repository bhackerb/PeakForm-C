"""MacroFactor XLSX multi-sheet parser.

Handles date normalization (Python datetime objects and M/D/YYYY strings),
numeric coercion, and returns tidy DataFrames per sheet.
"""

from __future__ import annotations

import re
from datetime import datetime, date
from typing import Dict, Optional

import pandas as pd
import openpyxl


# ---------------------------------------------------------------------------
# Sheet name constants (as exported by MacroFactor)
# ---------------------------------------------------------------------------

SHEET_CALORIES_MACROS = "Calories & Macros"
SHEET_SCALE_WEIGHT = "Scale Weight"
SHEET_WEIGHT_TREND = "Weight Trend"
SHEET_EXPENDITURE = "Expenditure"
SHEET_NUTRITION_PROGRAM = "Nutrition Program Settings"
SHEET_MICRONUTRIENTS = "Micronutrients"
SHEET_MUSCLE_GROUPS = "Muscle Groups - Sets"
SHEET_EXERCISES_VOLUME = "Exercises - Total Volume"
SHEET_EXERCISES_HEAVIEST = "Exercises - Heaviest Weight"

PRIORITY_SHEETS = [
    SHEET_CALORIES_MACROS,
    SHEET_SCALE_WEIGHT,
    SHEET_WEIGHT_TREND,
    SHEET_EXPENDITURE,
    SHEET_NUTRITION_PROGRAM,
    SHEET_MICRONUTRIENTS,
    SHEET_MUSCLE_GROUPS,
    SHEET_EXERCISES_VOLUME,
    SHEET_EXERCISES_HEAVIEST,
]


# ---------------------------------------------------------------------------
# Date normalization helpers
# ---------------------------------------------------------------------------

_MDYYYY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _normalize_date(value) -> Optional[date]:
    """Convert a MacroFactor date cell value to a Python date object.

    MacroFactor exports dates as either:
      - Python datetime / date objects (openpyxl reads them natively)
      - Strings in M/D/YYYY format
    Returns None for unrecognisable values.
    """
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        m = _MDYYYY_RE.match(value.strip())
        if m:
            month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(year, month, day)
        # Try ISO format as fallback
        try:
            return date.fromisoformat(value.strip()[:10])
        except ValueError:
            return None
    # openpyxl sometimes returns floats for Excel date serials — handle via pandas
    return None


def _to_float(value) -> Optional[float]:
    """Coerce a cell value to float, returning None on failure."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Generic sheet loader
# ---------------------------------------------------------------------------

def _load_sheet_as_df(wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
    """Load a worksheet into a DataFrame with normalised dates in column 0."""
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    data = list(ws.values)
    if not data:
        return pd.DataFrame()

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
    n = len(headers)
    rows = []
    for row in data[1:]:
        r = list(row)
        # openpyxl omits trailing None cells, so row length can differ from header length
        rows.append((r + [None] * n)[:n])

    df = pd.DataFrame(rows, columns=headers)

    # Normalise the first column as a date (it is always the date column)
    date_col = headers[0]
    df[date_col] = df[date_col].apply(_normalize_date)
    df = df.dropna(subset=[date_col])
    df = df.rename(columns={date_col: "date"})
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)

    return df


# ---------------------------------------------------------------------------
# Per-sheet parsers
# ---------------------------------------------------------------------------

def _parse_calories_macros(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_CALORIES_MACROS)
    if df.empty:
        return df
    numeric_cols = [c for c in df.columns if c != "date"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    # Normalise column names to lowercase snake_case
    df.columns = [
        c.lower().replace(" ", "_").replace("(", "").replace(")", "")
        if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_scale_weight(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_SCALE_WEIGHT)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.columns = [
        c.lower().replace(" ", "_").replace("%", "pct")
        if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_weight_trend(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_WEIGHT_TREND)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.columns = [
        c.lower().replace(" ", "_") if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_expenditure(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_EXPENDITURE)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.columns = [
        c.lower().replace(" ", "_") if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_nutrition_program(wb: openpyxl.Workbook) -> pd.DataFrame:
    """Nutrition Program Settings — contains the active calorie/macro targets."""
    df = _load_sheet_as_df(wb, SHEET_NUTRITION_PROGRAM)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.columns = [
        c.lower().replace(" ", "_").replace("(", "").replace(")", "")
        if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_micronutrients(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_MICRONUTRIENTS)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.columns = [
        c.lower().replace(" ", "_").replace("(", "").replace(")", "")
        .replace("/", "_per_").replace("-", "_")
        if c != "date" else "date"
        for c in df.columns
    ]
    return df


def _parse_muscle_groups(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_MUSCLE_GROUPS)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _parse_exercises_volume(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_EXERCISES_VOLUME)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _parse_exercises_heaviest(wb: openpyxl.Workbook) -> pd.DataFrame:
    df = _load_sheet_as_df(wb, SHEET_EXERCISES_HEAVIEST)
    if df.empty:
        return df
    for col in [c for c in df.columns if c != "date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class MacroFactorData:
    """Container for all parsed MacroFactor sheets."""

    def __init__(self, filepath: str):
        self._filepath = filepath
        self.wb: openpyxl.Workbook = openpyxl.load_workbook(
            filepath, data_only=True, read_only=True
        )
        self._sheets: Dict[str, pd.DataFrame] = {}

    def _get(self, sheet_name: str, parser_fn) -> pd.DataFrame:
        if sheet_name not in self._sheets:
            self._sheets[sheet_name] = parser_fn(self.wb)
        return self._sheets[sheet_name]

    @property
    def calories_macros(self) -> pd.DataFrame:
        return self._get(SHEET_CALORIES_MACROS, _parse_calories_macros)

    @property
    def scale_weight(self) -> pd.DataFrame:
        return self._get(SHEET_SCALE_WEIGHT, _parse_scale_weight)

    @property
    def weight_trend(self) -> pd.DataFrame:
        return self._get(SHEET_WEIGHT_TREND, _parse_weight_trend)

    @property
    def expenditure(self) -> pd.DataFrame:
        return self._get(SHEET_EXPENDITURE, _parse_expenditure)

    @property
    def nutrition_program(self) -> pd.DataFrame:
        return self._get(SHEET_NUTRITION_PROGRAM, _parse_nutrition_program)

    @property
    def micronutrients(self) -> pd.DataFrame:
        return self._get(SHEET_MICRONUTRIENTS, _parse_micronutrients)

    @property
    def muscle_groups(self) -> pd.DataFrame:
        return self._get(SHEET_MUSCLE_GROUPS, _parse_muscle_groups)

    @property
    def exercises_volume(self) -> pd.DataFrame:
        return self._get(SHEET_EXERCISES_VOLUME, _parse_exercises_volume)

    @property
    def exercises_heaviest(self) -> pd.DataFrame:
        return self._get(SHEET_EXERCISES_HEAVIEST, _parse_exercises_heaviest)

    def available_sheets(self):
        return self.wb.sheetnames

    def get_current_targets(self) -> Dict[str, Optional[float]]:
        """Return the most recent active nutrition targets from MacroFactor.

        Looks at Nutrition Program Settings and returns the last row's values.
        Falls back to config defaults if the sheet is unavailable.
        """
        from peakform.config import (
            FALLBACK_CALORIE_TARGET,
            FALLBACK_PROTEIN_TARGET_G,
            FALLBACK_CARBS_TARGET_G,
            FALLBACK_FAT_TARGET_G,
            FALLBACK_EXPENDITURE_KCAL,
        )

        df = self.nutrition_program
        defaults = {
            "calories": FALLBACK_CALORIE_TARGET,
            "protein_g": FALLBACK_PROTEIN_TARGET_G,
            "carbs_g": FALLBACK_CARBS_TARGET_G,
            "fat_g": FALLBACK_FAT_TARGET_G,
            "expenditure_kcal": FALLBACK_EXPENDITURE_KCAL,
        }
        if df.empty:
            return defaults

        last = df.iloc[-1]

        # Column names vary — try common patterns
        def _find_col(df: pd.DataFrame, *candidates) -> Optional[str]:
            for c in candidates:
                for col in df.columns:
                    if c.lower() in col.lower():
                        return col
            return None

        cal_col = _find_col(df, "calorie", "kcal", "energy")
        prot_col = _find_col(df, "protein")
        carb_col = _find_col(df, "carb")
        fat_col = _find_col(df, "fat")
        exp_col = _find_col(df, "expenditure")

        result = dict(defaults)
        if cal_col:
            v = _to_float(last[cal_col])
            if v:
                result["calories"] = v
        if prot_col:
            v = _to_float(last[prot_col])
            if v:
                result["protein_g"] = v
        if carb_col:
            v = _to_float(last[carb_col])
            if v:
                result["carbs_g"] = v
        if fat_col:
            v = _to_float(last[fat_col])
            if v:
                result["fat_g"] = v
        if exp_col:
            v = _to_float(last[exp_col])
            if v:
                result["expenditure_kcal"] = v

        return result


def load(filepath: str) -> MacroFactorData:
    """Load and parse a MacroFactor XLSX export."""
    return MacroFactorData(filepath)
